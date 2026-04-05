import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_FILE = os.path.join(BASE_DIR, "registration.xlsx")
SHEET_NAME = "Registrations"


HEADERS = [
    "Registered At (UTC)",
    "Full Name",
    "Email",
    "Phone",
    "Role",
]


def ensure_registration_workbook():
    if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
        try:
            wb = load_workbook(EXCEL_FILE)
            if SHEET_NAME not in wb.sheetnames:
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(HEADERS)
                wb.save(EXCEL_FILE)
            return
        except Exception:
            # If file is corrupt/empty-but-present, recreate it cleanly.
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)

def save_registration_to_excel(full_name, email, phone, role):
    ensure_registration_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append(
        [
            datetime.now(timezone.utc).isoformat(),
            full_name,
            email,
            phone,
            role,
        ]
    )
    wb.save(EXCEL_FILE)


def read_registrations_as_json():
    ensure_registration_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = r[i] if i < len(r) else None
        out.append(item)
    return out