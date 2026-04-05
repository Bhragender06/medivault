import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook


SHEET_NAME = "Registrations"
HEADERS = ["Registered At (UTC)", "Full Name", "Email", "Phone", "Role"]


def _abs_path(config_path: str) -> str:
    # Accept either absolute paths or repo-relative paths.
    if os.path.isabs(config_path):
        return config_path
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    return os.path.abspath(os.path.join(repo_root, config_path))


def ensure_workbook(xlsx_path: str) -> str:
    path = _abs_path(xlsx_path)
    os.makedirs(os.path.dirname(path), exist_ok=True)

    if os.path.exists(path) and os.path.getsize(path) > 0:
        try:
            wb = load_workbook(path)
            if SHEET_NAME not in wb.sheetnames:
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(HEADERS)
                wb.save(path)
            return path
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(path)
    return path


def append_registration(xlsx_path: str, full_name: str | None, email: str, phone: str | None, role: str):
    path = ensure_workbook(xlsx_path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    ws.append([datetime.now(timezone.utc).isoformat(), full_name, email, phone, role])
    wb.save(path)


def registrations_as_json(xlsx_path: str) -> list[dict]:
    path = ensure_workbook(xlsx_path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = r[i] if i < len(r) else None
        out.append(item)
    return out

